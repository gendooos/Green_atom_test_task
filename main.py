from selenium.webdriver.common.by import By
from config import driver, wait
import time
from datetime import datetime
import calendar
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import json

def main():
    #1. Открыть https://www.moex.com;
    driver.get(r'https://www.moex.com/')

    # 2. Перейти по следующим элементам: Меню -> Срочный рынок -> Индикативные курсы;
    driver.find_element(By.XPATH,"//a[@class='fast-links__link' and contains(text(),'Срочный рынок')]").click()
    driver.switch_to.window(driver.window_handles[1])
    wait.until(lambda d:driver.find_element(By.XPATH,"//a[@class='btn2 btn2-primary' and contains(text(),'Согласен')]"))
    driver.find_element(By.XPATH,"//a[@class='btn2 btn2-primary' and contains(text(),'Согласен')]").click()
    wait.until(lambda d: driver.find_element(By.XPATH,"//span[contains(text(),'Индикативные курсы')]").is_displayed())
    driver.find_element(By.XPATH,"//span[contains(text(),'Индикативные курсы')]").click()

    # 3. В выпадающем списке выбрать валюты: USD/RUB - Доллар США к российскому рублю;
    last_month_currency('USD/RUB - Доллар США к российскому рублю')

    # 4. Сформировать данные за предыдущий месяц;
    driver.find_element(By.XPATH, '//button[@aria-label="Показать"]').click()

    # 5. Скопировать данные в Excel;
    wb = openpyxl.Workbook()
    wb.save("greenatom_test.xlsx")
    wb = openpyxl.load_workbook('greenatom_test.xlsx')
    wb_list = wb[wb.sheetnames[0]]
    wait.until(lambda d: driver.find_element(By.XPATH,'//tr[@class="ui-table-row -interactive"]').is_displayed())
    copy_to_excel(wb, wb_list,'Дата USD/RUB', 'Курс USD/RUB', 'Время USD/RUB')

    # 6. Повторить шаги для валют JPY/RUB - Японская йена к российскому рублю;
    last_month_currency('JPY/RUB - Японская йена к российскому рублю')
    driver.find_element(By.XPATH, '//button[@aria-label="Показать"]').click()

    # 7. Скопировать данные в Excel;
    wait.until(lambda d: driver.find_element(By.XPATH,'//tr[@class="ui-table-row -interactive"]').is_displayed())
    copy_to_excel(wb, wb_list,'Дата JPY/RUB', 'Курс JPY/RUB', 'Время JPY/RUB')

    # 8. Для каждой строки полученного файла поделить курс USD/RUB на JPY/RUB, полученное значение записать в ячейку (G) Результат
    div_calc(wb, wb_list)

    # 9. Выровнять – автоширина;
    set_auto_width(wb_list)
    wb.save("greenatom_test.xlsx")

    # 12. Направить итоговый файл отчета себе на почту;
    send_email('greenatom_test.xlsx', wb_list.max_row)


def last_month_currency(currency):
    driver.find_element(By.XPATH,"//span[@class = 'ui-icon ui-select__icon -arrow']").click()
    driver.find_element(By.XPATH,f"//a[contains(text(),'{currency}')]").click()
    driver.find_element(By.XPATH,"//label[@for = 'fromDate']").click()
    wait.until(lambda d:driver.find_element(By.XPATH,'//div[@class = "ui-dropdown ui-calendar -opened"]').is_displayed())
    prev_month = driver.find_element(By.XPATH,'//div[@class = "ui-dropdown ui-calendar -opened"]').find_element(By.CLASS_NAME, 'ui-select__placeholder').text[:2]
    choose_perv_month(prev_month)
    driver.find_element(By.XPATH,"//label[@for = 'tillDate']").click()
    wait.until(lambda d: driver.find_element(By.XPATH,'//div[@class = "ui-dropdown ui-calendar -opened"]').find_element(By.XPATH, './/div[@class = "ui-select__activator -selected"]').is_displayed())
    choose_perv_month(prev_month, calendar.monthrange(datetime.now().year,datetime.now().month - 1)[1])
    
def choose_perv_month(prev_month, day = '1'):
    driver.find_element(
        By.XPATH,'//div[@class = "ui-dropdown ui-calendar -opened"]'
                        ).find_element(
                            By.XPATH, './/div[@class = "ui-select__activator -selected"]'
                            ).click()
    driver.find_element(
        By.XPATH,'//div[@class = "ui-dropdown ui-calendar__dropdown -opened"]'
                        ).find_element(
                            By.XPATH, f'.//div[@class = "ui-select-option__content" and contains(text(),"{prev_month}")]').click()
    driver.find_element(
        By.XPATH,'//div[@class = "ui-dropdown ui-calendar -opened"]'
                        ).find_element(
                            By.XPATH,f".//div[@class='ui-calendar__cell -day' and contains(text(),'{day}')]").click()

def copy_to_excel(wb, wb_list, first,second,third):
    # wb = openpyxl.load_workbook('greenatom_test.xlsx')
    # wb_list = wb[wb.sheetnames[0]]
    if wb_list.cell(1,1).value == None:
        wb_list.cell(1,1).value = first
    else:
        wb_list.cell(1,wb_list.max_column+1).value = first
    wb_list.cell(1,wb_list.max_column+1).value = second
    wb_list.cell(1,wb_list.max_column+1).value = third
    rows = driver.find_elements(By.XPATH,'//tr[@class="ui-table-row -interactive"]')
    print(rows)
    min_row_count = 2
    for web_row in rows:
        web_cells = web_row.find_elements(By.XPATH,'.//*')
        date = web_cells[0].text
        value = web_cells[3].text
        date_time = web_cells[4].text
        wb_list.cell(min_row_count, wb_list.max_column-2).value = date
        print(wb_list.cell(min_row_count, wb_list.max_column-2).value, min_row_count, wb_list.max_column-2 )
        wb_list.cell(min_row_count, wb_list.max_column-1).value = value
        wb_list.cell(min_row_count, wb_list.max_column).value = date_time
        min_row_count += 1
    wb.save("greenatom_test.xlsx")

def div_calc(wb, wb_list):
    wb_list[f'G1'].value = 'Деление USD/RUB на JPY/RUB'
    for row in wb_list.iter_rows(min_row=2):
        # 10. Формат чисел – финансовый;
        wb_list[f'B{row[0].row}'].number_format = '"₽"#,##0.00'
        wb_list[f'E{row[0].row}'].number_format = '"₽"#,##0.00'
        wb_list[f'G{row[0].row}'].value = float(wb_list[f'B{row[0].row}'].value) / float(wb_list[f'E{row[0].row}'].value)
    wb.save("greenatom_test.xlsx")
    
def set_auto_width(wb_list):
    for column in wb_list.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                print(e)
        adjusted_width = (max_length + 2)
        wb_list.column_dimensions[column_letter].width = adjusted_width

def send_email(attachment, row_count):
    credentials = load_email_credentials('personal_date.json')
    subject = 'greenatom_test.xlsx'
    # 13. В письме указать количество строк в Excel в правильном склонении.
    body = f"Количество строк в отчете: {row_count} {decline_row_count(row_count)}"
    
    msg = MIMEMultipart()
    msg['From'] = credentials["email"]
    msg['To'] = credentials["email"]
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    with open(attachment, 'rb') as file:
        part = MIMEApplication(file.read())
        part.add_header('Content-Disposition', f'attachment; filename="{attachment}"')
        msg.attach(part)

    with smtplib.SMTP(credentials["smtp"], 587) as server:
        server.starttls()
        server.login(credentials["email"], credentials["pass"])
        server.send_message(msg)

def load_email_credentials(filepath):
    with open(filepath, 'r') as file:
        return json.load(file)

def decline_row_count(count):
    if count % 10 == 1 and count % 100 != 11:
        return 'строка'
    elif 2 <= count % 10 <= 4 and not (12 <= count % 100 <= 14):
        return 'строки'
    else:
        return 'строк'

if __name__ == '__main__':
    main()